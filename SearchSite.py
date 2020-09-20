# This is a script to search up ACU KWIK and fill up Table 1.
from bs4 import BeautifulSoup
# from tkinter import *
# import tkinter.filedialog as tk
# import tkinter.messagebox as tk2
import os
import urllib.request, urllib.parse
import pandas as pd
import numpy as np
import time
from pathlib import Path


# Function to grab important data from list of text and turns into pandas dataframe object
def grabimpdatacol():
    coldata = []
    # 1. Location column
    coldata.append('LOCATION')
    # 2. ICAO&IATA columns
    coldata.append('ICAO')
    coldata.append('IATA')
    # 3. Elevation column
    coldata.append('ELEVATION')
    # 4. Runway length and width
    coldata.append('RUNWAY LENGTH')
    coldata.append('WIDTH')
    # 5. Runway surface type
    coldata.append('RUNWAY SURFACE')
    # Statement below prints for columns of dataframe
    # print(coldata)
    return coldata


def grabmiscdatacol():
    coldata = []
    # 1. ICAO code
    coldata.append('ICAO')
    # 2. Latitude
    coldata.append('LATITUDE')
    # 3. Longitude
    coldata.append('LONGITUDE')
    # 4. Airport type
    coldata.append('AIRPORT TYPE')
    # 5. Airport of entry
    coldata.append('AIRPORT OF ENTRY')
    # 6. Customs
    coldata.append('CUSTOMS')
    # 7. US Preclearance
    coldata.append('US-CUSTOMS PRECLEARANCE')
    # 8. Open 24H
    coldata.append('OPEN 24H')
    # 9. Time per url parse
    coldata.append('Average time per URL parse')

    # Statement below prints for columns of miscellaenous dataframe
    # print(coldata)
    return coldata


def grabmiscdataval(mystr, checkvalue=1):
    valdata = []
    # 1. ICAO code
    valdata.append(mystr[1][7:11])
    # 2. Latitude
    posslash = mystr[3].find('/')
    valdata.append(mystr[3][:posslash])
    # 3. Longitude
    valdata.append(mystr[3][posslash + 1:])
    # 4. Airport type
    for i in range(len(mystr)):
        if mystr[i] == 'Airport Type':
            valdata.append(mystr[i + 1])
    # 5. Airport of Entry
    for i in range(len(mystr)):
        if mystr[i] == 'Airport of Entry':
            valdata.append(mystr[i + 1])
    # 6. Customs
    for i in range(len(mystr)):
        if mystr[i] == 'Customs':
            valdata.append(mystr[i + 1])
    # 7. US Customs Pre-clearance
    for i in range(len(mystr)):
        if mystr[i] == 'US Customs Pre-Clearance':
            valdata.append(mystr[i + 1])
    # 8. Open 24 hours
    for i in range(len(mystr)):
        if mystr[i] == 'Open 24 Hours':
            valdata.append(mystr[i + 1])

    # Below statement prints all miscellaneous values of dataframe
    # print(valdata)
    return valdata


def grabimpdataval(mystr):
    '''
    DOCSTRING: A function that returns the data entry values into table 1 for a particular ICAO code.
    This function is called in getdata(), which in turn is looped in main()
    input: The parts of a string obtained by parsing through html code of a particular url with given ICAO code
    outut: The data entry values for Table 1 for that ICAO code
    '''

    valdata = []
    # 1. Location
    valdata.append(mystr[0][11:])
    # 2. ICAO and IATA
    valdata.append(mystr[1][7:11])
    valdata.append(mystr[1][20:23])
    # 3. Elevation
    valdata.append(mystr[5])
    # 4. Runway length and width
    posx = mystr[11].find('x')
    valdata.append(mystr[11][:posx - 1])
    poscomma = mystr[11].find(',')
    valdata.append(mystr[11][posx + 2:poscomma])
    # 5. Runway surface type
    valdata.append(mystr[13])
    # Below statement prints all important values of dataframe
    # print(valdata)
    return valdata


def dataframebuilder(coldata, valdata):
    df = pd.DataFrame(data=valdata, columns=coldata)
    return df

    # df.to_excel("D:/Semesters/2019-20/Summer 2020/Full-time/TestRun.xls")


# Function to search ICAO code with given airport name
def searchICAO(apname):
    # returns ICAO code
    pass


def getdata(ICAOcode):
    '''
    DOCSTRING: Function to get data by parsing through html code on the website by using ICAO code as an identifier in the url
    input: ICAO code
    output: Data entry values by calling the grabimpdataval(). getdata() is called through main() and looped through for different ICAO codes.
    '''
    url = 'https://acukwik.com/Airport-Info/' + ICAOcode

    response = urllib.request.urlopen(url)
    # print(response)
    webContent = response.read()
    # print(webContent)

    soup = BeautifulSoup(webContent, 'html.parser')
    # print(soup.prettify())
    # text = soup.get_text()

    i = 0
    importantstr = []
    liststring = soup.stripped_strings

    for string in liststring:
        i += 1
        if string.startswith(ICAOcode.upper() + ' -'):
            count = i
            break

    for string in liststring:
        if i >= count and i <= count + 55:
            importantstr.append(string)
        elif i > count + 55:
            break
        i += 1

    '''
    i = 0
    importantstr = []
    for string in soup.stripped_strings:
        i += 1
        if i >= 19 and i <= 73:
            importantstr.append(string)
        elif i > 73:
            break
    '''

    # Statement below checks for parse
    # print(importantstr)

    # grabbing values of eventual dataframe
    return (grabimpdataval(importantstr), grabmiscdataval(importantstr))


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, truncate_sheet=False, **to_excel_kwargs):
    '''
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    '''
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def excelwriter(df, checkval=0):
    mystr = os.path.dirname(os.path.realpath(__file__))
    mystr = Path(mystr)
    mystr = mystr.parent.parent.parent

    if checkval == 0:
        with pd.ExcelWriter(str(mystr) + "\Part1\Output\Source.xlsx", engine="openpyxl",
                            mode='w') as writer:
            df.to_excel(writer, sheet_name='DataEntry')
    elif checkval == 1:
        with pd.ExcelWriter(str(mystr) + "\Part1\Output\Source.xlsx", engine="openpyxl",
                            mode='a') as writer:
            df.to_excel(writer, sheet_name='MiscEntry')
    else:
        # database collector!!
        # append_df_to_excel(str(mystr) + "\Output\Database.xlsx", df, sheet_name='Sheet1', startrow=None,
        # truncate_sheet=False, header=False, index=False)

        bombardierIDs = []
        mystr = os.path.dirname(os.path.realpath(__file__))
        mystr = Path(mystr)
        mystr = mystr.parent.parent.parent

        f = open(str(mystr) + "\Part1\Bombardier ID.txt", "r")
        lines = f.readlines()

        for i in range(len(lines)):
            if ':' in lines[i]:
                idstring = lines[i][lines[i].find(':') + 1:-1]
                bombardierIDs.append(idstring.strip())
        f.close()

        # C:\\Users\IDNUMBER\Bombardier\Sales Engineering - Shared Database Test\Database.xlsx
        filepathlist = [
            'Bombardier\Sales Engineering - Documents\Sales Engineering - DO NOT SYNC\Performance Tools\The Dream\Synced Files',
            'Bombardier\Sales Engineering - Sales Engineering - DO NOT SYNC\Performance Tools\The Dream\Synced Files',
            'Bombardier\Sales Engineering - Performance Tools\The Dream\Synced Files',
            'Bombardier\Sales Engineering - The Dream\Synced Files', 'Bombardier\Sales Engineering - Synced Files']

        for i in range(len(bombardierIDs)):
            for j in range(len(filepathlist)):
                try:
                    append_df_to_excel(
                        f"C:\\Users\\{bombardierIDs[i]}\\{filepathlist[j]}\Database.xlsx",
                        df, sheet_name='Sheet1', startrow=None,
                        truncate_sheet=False, header=False, index=False)
                except:
                    pass

        # load workbook, remove duplicate ICAO and re-write

        # loadeddf = pd.read_excel(str(mystr) + "\Output\Database.xlsx")
        for i in range(len(bombardierIDs)):
            for j in range(len(filepathlist)):
                try:
                    loadeddf = pd.read_excel(
                        f"C:\\Users\\{bombardierIDs[i]}\\{filepathlist[j]}\Database.xlsx")
                    loadeddf.drop_duplicates(inplace=True)
                except:
                    pass

        # with pd.ExcelWriter(str(mystr) + "\Output\Database.xlsx", engine="openpyxl",
        # mode='w') as writer:
        for i in range(len(bombardierIDs)):
            for j in range(len(filepathlist)):
                try:
                    with pd.ExcelWriter(
                            f"C:\\Users\\{bombardierIDs[i]}\\{filepathlist[j]}\Database.xlsx",
                            engine="openpyxl", mode='w') as writer:
                        loadeddf.to_excel(writer, sheet_name='Sheet1', index=False)
                except:
                    pass


'''
def guibuilder():

    ICAOlist = []

    class Application(Frame):

        def __init__(self, master):
            super(Application, self).__init__(master)

            # self.create_widgets()
            self.ICAOcodebox = Listbox(self, width=40, height=10,
                                       selectmode=SINGLE)

            for code in ICAOlist:
                self.ICAOcodebox.insert(END, code)

            self.grid(rowspan=5, columnspan=4)
            self.ICAOcodebox.grid(row=1)
            self.generateButton = Button(self, text='Generate', command=self.generate)
            self.generateButton.grid(row=4, column=0)
            self.pack()

        def generate(self):
            if (len(ICAOlist) == 0):
                tk2.showinfo('Notice', 'No ICAO codes entered yet!\nClick Add to add ICAO codes.')
            else:
                pass
            #SOMETHING HERE

    root = Tk()
    root.title('Quick Data Entry')
    root.geometry('500x200')
    e = Entry(root, width=50)
    e.pack()
    e.insert(0,'Default code - CYUL')

    def myClick():
        mylabel = Label(root,e.get())
        mylabel.pack()

    myButton = Button(root,text = 'Enter an ICAO code',command = myClick)
    myButton.pack()
    root.mainloop()
    #app = Application(root)
    #app.mainloop()
'''


def myinput():
    codelist = []
    ans = 'Y'
    code = ""
    print('\nEnter the ICAO codes or the Cities of desired airports.\nType ESCAPE to end the list.\n')
    i = 1
    while code.lower() != 'escape':
        code = input(f'{i}. ')
        codelist.append(code)
        i += 1
        # ans = input('Enter another code/city? Y for yes, otherwise for no\n')
    codelist.pop()
    return codelist


# Main wrapper function
def main():
    # CHOICE 1: RESUME BELOW LINE
    codelist = myinput()

    # CHOICE 2: RESUME BELOW LINE
    # codelist = guibuilder()

    # ICAOcode = ['WSSS','KLAX','KASE','CYUL','EGLC','KJFK','LTFM','OOMS','CYWG','LIRQ','LIRF','LSZH','LFPG','ZKPY','FVRG']

    start_time = time.time()
    urlsumtime = 0
    # grabbing column names of eventual dataframe
    coldata = grabimpdatacol()
    misccoldata = grabmiscdatacol()
    # The ICAO codes are parsed through for the important data in loops
    valdata = []
    newvaldata = []
    miscdata = []

    i = 0
    for codes in codelist:
        try:
            urlstarttime = time.time()
            valdata.append(getdata(codes))
            urlsumtime = urlsumtime + (time.time() - urlstarttime)
            i += 1

        except:
            # INCLUDE airport comment
            # mystr = os.path.dirname(os.path.realpath(__file__))
            # mystr = Path(mystr)
            # mystr = mystr.parent.parent
            # loadeddf = pd.read_excel(str(mystr) + "\Output\Database.xlsx")
            bombardierIDs = []
            mystr = os.path.dirname(os.path.realpath(__file__))
            mystr = Path(mystr)
            mystr = mystr.parent.parent.parent
            f = open(str(mystr) + "\Part1\Bombardier ID.txt", "r")
            lines = f.readlines()

            for lv1 in range(len(lines)):
                if ':' in lines[lv1]:
                    idstring = lines[lv1][lines[lv1].find(':') + 1:-1]
                    bombardierIDs.append(idstring.strip())

            f.close()

            # C:\\Users\IDNUMBER\Bombardier\Sales Engineering - Shared Database Test\Database.xlsx
            filepathlist = [
                'Bombardier\Sales Engineering - Documents\Sales Engineering - DO NOT SYNC\Performance Tools\The Dream\Synced Files',
                'Bombardier\Sales Engineering - Sales Engineering - DO NOT SYNC\Performance Tools\The Dream\Synced Files',
                'Bombardier\Sales Engineering - Performance Tools\The Dream\Synced Files',
                'Bombardier\Sales Engineering - The Dream\Synced Files', 'Bombardier\Sales Engineering - Synced Files']

            for lv1 in range(len(bombardierIDs)):
                for lv2 in range(len(filepathlist)):
                    try:
                        loadeddf = pd.read_excel(
                            f"C:\\Users\\{bombardierIDs[lv1]}\\{filepathlist[lv2]}\Database.xlsx")
                    except:
                        pass

            loadedseries = loadeddf['City'].str.lower()
            mycitycode = loadeddf[loadedseries.str.contains(codes.lower(), regex=False)]['ICAO'].values

            # print(mycitycode)
            # print(type(mycitycode))

            if len(mycitycode) > 0:
                print(f"\nThe corresponding ICAO code(s) for '{codes}' listed on the database include:")
                for j in range(len(mycitycode)):
                    print(f'{j + 1}. {mycitycode[j]}\t\t', end="")
                print('\n')
                while True:
                    try:
                        while True:
                            mynuminput = int(input('Enter the corresponding number.\t'))
                            if mynuminput >= 1 and mynuminput <= len(mycitycode):
                                urlstarttime = time.time()
                                valdata.append(getdata(mycitycode[mynuminput - 1]))
                                urlsumtime = urlsumtime + (time.time() - urlstarttime)
                                i += 1
                                break
                            else:
                                print('Please enter one of the numbers displayed above.\n')
                                continue
                    except TypeError:
                        print('Type error occurred... Make sure you are typing a number!')
                        continue
                    except:
                        print('Code not found on ACU-KWIK... Try another number!')
                        continue
                    break

            else:
                # print('I am here')
                # print(i)
                # print(codelist)

                print(
                    f"'{codes.upper()}' ICAO code does not exist on the ACU-KWIK website OR Airport city is not present in database.")
                codelist[i] = 'ERROR'
                i += 1

    # grabbing miscellaneous data from valdata and redefining valdata
    for i in range(len(valdata)):
        miscdata.append(valdata[i][1])
        newvaldata.append(valdata[i][0])

    print(valdata)
    print('\n\n')
    print(newvaldata)
    print('\n\n')
    print(miscdata)
    print('\n\n')

    codelist = list(filter(lambda mystr: mystr != 'ERROR', codelist))

    print(codelist)
    print('\n\n')

    try:
        # 9.Avg time per url parse
        add1 = (urlsumtime / len(codelist))
        add2 = 0
        check = 0
        firsttime = True
        flattenlist = sum(miscdata, [])

        print('Fail 1')

        for i in range(1, len(codelist) * 8 + 1 + check):
            if i % 8 == 0:
                if firsttime == True:
                    flattenlist.insert(i, add1)
                    firsttime = False
                    check += 1
                else:
                    flattenlist.insert(i + check, add2)
                    check += 1

        print('Fail 2')

        # redefine miscellaneous data
        miscdata = flattenlist
        newvaldata = np.array(newvaldata).reshape(len(codelist), 7)
        miscdata = np.array(miscdata).reshape(len(codelist), 9)

        print('\n\n')
        print(newvaldata)
        print(miscdata)

        print('Fail 3')

        df = dataframebuilder(coldata, newvaldata)
        excelwriter(df)

        print('Fail 4')

        # checkval argument to denote that miscellaneous dataframe is being built
        df = dataframebuilder(misccoldata, miscdata)
        excelwriter(df, checkval=1)

        print('Fail 5')

        # dataframe for database being built. Database will have only ICAO code and city name
        databasedata = []

        for i in range(len(codelist)):
            poscomma = newvaldata[i][0].find(',')
            cityname = newvaldata[i][0][:poscomma]
            databasedata.append(newvaldata[i][1])
            databasedata.append(cityname)
        databasedata = np.array(databasedata).reshape(len(codelist), 2)
        databasecoldata = coldata[:2]

        print('Fail 6')
        print(databasedata)
        print('\n\n')
        print(databasecoldata)

        df = dataframebuilder(databasecoldata[::-1], databasedata)
        excelwriter(df, checkval=2)

        print('Fail 7')

        # elevation = newvaldata[:, 3]

        topasscode = newvaldata[:, 1]

        print(topasscode)
        print('Fail 8')

        print("\n--- %s seconds (excluding initial list input) ---\n" % (time.time() - start_time))

        print('Fail 9')

        return topasscode

    except:
        print("No valid ICAO codes entered and/or no Airport city names present in database!")
        return 0


def tempforAPG(airportcode=['NA']):
    # Temperatures directly passed in!
    mystr = os.path.dirname(os.path.realpath(__file__))
    mystr = Path(mystr)
    mystr = mystr.parent.parent.parent

    loadeddf = pd.read_excel(str(mystr) + "\Part1\Output\Source.xlsx")
    temperatures = loadeddf.iloc[:, -1]
    temperatureSeries = pd.Series(temperatures)
    temperatureSeries.fillna('MISSING', inplace=True)

    for i in range(len(temperatureSeries)):
        if (type(temperatureSeries[i]) != np.float64 and type(temperatureSeries[i]) != np.int64):
            return 0

    if len(temperatureSeries) != len(airportcode):
        return 1
    return temperatureSeries.values


def grooveforAPG():
    mystr = os.path.dirname(os.path.realpath(__file__))
    mystr = Path(mystr)
    mystr = mystr.parent.parent.parent

    loadeddf = pd.read_excel(str(mystr) + "\Part1\Output\Source.xlsx")
    dfgrooved = loadeddf[loadeddf['RUNWAY SURFACE'].apply(lambda mystr: True if 'groove' in mystr.lower() else False)]
    dfgrooved = dfgrooved.iloc[:, np.r_[2, -1]]

    dfnongrooved = loadeddf[
        loadeddf['RUNWAY SURFACE'].apply(lambda mystr: False if 'groove' in mystr.lower() else True)]
    dfnongrooved = dfnongrooved.iloc[:, np.r_[2, -1]]

    grooveicaocodelist = dfgrooved['ICAO']
    tempgroovelist = dfgrooved.iloc[:, -1]
    nongrooveicaocodelist = dfnongrooved['ICAO']
    tempnongroovelist = dfnongrooved.iloc[:, -1]

    return (grooveicaocodelist.values.tolist(), tempgroovelist.values.tolist(), nongrooveicaocodelist.values.tolist(),
            tempnongroovelist.values.tolist())


def ICAOcodeforAPG():
    # ICAO codes directly passed in
    mystr = os.path.dirname(os.path.realpath(__file__))
    mystr = Path(mystr)
    mystr = mystr.parent.parent.parent

    loadeddf = pd.read_excel(str(mystr) + "\Part1\Output\Source.xlsx")
    icaocodes = loadeddf['ICAO']

    '''
    if type(icaocodes[0]) != np.float64 and type(icaocodes[0]) != np.int64:
        return 0
    elif len(temperatureSeries) != len(airportcode):
        return 1
    '''
    return icaocodes.values.tolist()
